"""Write rows to .xlsx matching the DDF schema (SKU as text, ordered columns)."""
import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .models import COLUMNS, ProductRow


def write_xlsx(rows: list[ProductRow], path: str) -> None:
    records = [r.as_ordered_dict() for r in rows]
    df = pd.DataFrame(records, columns=COLUMNS)
    # SKU must stay text so leading zeros / non-numeric ids survive.
    df["TIRA_SKU"] = df["TIRA_SKU"].astype("string")

    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        df.to_excel(xw, index=False, sheet_name="Sheet1")
        ws = xw.sheets["Sheet1"]
        for col_idx, name in enumerate(COLUMNS, start=1):
            ws.cell(row=1, column=col_idx).font = Font(bold=True)
            width = max(len(name), 18)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 60)
